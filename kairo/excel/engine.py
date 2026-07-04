# PROVENANCE: original | clean-room Excel domain engine per prompts/domains/08c_excel.md + prompts/07_excel_real_calculation.md
"""Excel domain engine — real formula computation via LibreOffice headless recompute.

Implements the ``xlsx_recompute`` and ``xlsx_structure_readback`` oracles from
specs/VERIFICATION_ORACLES.md for the Excel domain.

ARCHITECTURE:
  1. openpyxl (MIT) writes formulas + structure to .xlsx files.
  2. LibreOffice headless (soffice --headless, MPL) recomputes all formula values
     as a local subprocess — NOT linked into our code, no AGPL introduced.
  3. Read recomputed cell VALUES back with openpyxl (data_only=True) and compare
     to an INDEPENDENT Python computation (numpy/pandas, BSD) of the same logic.
  4. Never trust the writer alone — the oracle verifies recomputed values vs
     independent calc.

HONEST DEGRADATION:
  If LibreOffice is not installed, the engine FAILS LOUD:
  "recompute engine unavailable — install LibreOffice"
  It NEVER presents unverified formula strings as done.

Dependencies:
  - openpyxl (MIT) — xlsx read/write
  - LibreOffice headless (MPL) — formula recompute subprocess
  - numpy (BSD-3) — independent calculation
  - cryptography (Apache-2.0/BSD-3) — Ed25519 audit + egress report

All operations are fully offline. No network calls. No LLM. No cloud.
"""

from __future__ import annotations

import hashlib
import logging
import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

log = logging.getLogger("kairo.excel")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_LIBREOFFICE_BIN = None  # cached path


def _find_libreoffice() -> str | None:
    """Find the LibreOffice binary on the system.

    Returns the path to the soffice/libreoffice binary, or None if not found.
    """
    global _LIBREOFFICE_BIN
    if _LIBREOFFICE_BIN is not None:
        return _LIBREOFFICE_BIN

    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            _LIBREOFFICE_BIN = path
            return path

    # Check common install locations
    common_paths = [
        "/usr/bin/soffice",
        "/usr/bin/libreoffice",
        "/usr/local/bin/soffice",
        "/opt/libreoffice/program/soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "C:\\Program Files\\LibreOffice\\program\\soffice.exe",
    ]
    for p in common_paths:
        if os.path.isfile(p):
            _LIBREOFFICE_BIN = p
            return p

    return None


class LibreOfficeUnavailableError(RuntimeError):
    """Raised when LibreOffice is not installed — honest degradation.

    The engine FAILS LOUD rather than presenting unverified formula strings.
    """


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ExcelEdit:
    """A single Excel edit applied to a workbook."""

    sheet: str
    cell: str
    old_value: Any
    new_value: Any
    formula: str
    rationale: str


@dataclass(frozen=True)
class FlaggedCell:
    """A cell that was flagged (not edited) during the Excel run."""

    sheet: str
    cell: str
    reason: str


@dataclass
class ExcelResult:
    """Structured result of the Excel pipeline run."""

    ok: bool
    output_path: str = ""
    applied_edits: list[ExcelEdit] = field(default_factory=list)
    flagged_cells: list[FlaggedCell] = field(default_factory=list)
    recomputed_values: dict[str, Any] = field(default_factory=dict)
    independent_calc: dict[str, Any] = field(default_factory=dict)
    recompute_verified: bool = False
    error: str = ""
    audit_log_json: str = ""
    egress_report_json: str = ""
    doc_hash: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "output_path": self.output_path,
            "applied_edits": [
                {
                    "sheet": e.sheet,
                    "cell": e.cell,
                    "old_value": str(e.old_value),
                    "new_value": str(e.new_value),
                    "formula": e.formula,
                    "rationale": e.rationale,
                }
                for e in self.applied_edits
            ],
            "flagged_cells": [
                {"sheet": f.sheet, "cell": f.cell, "reason": f.reason}
                for f in self.flagged_cells
            ],
            "recompute_verified": self.recompute_verified,
            "error": self.error,
            "doc_hash": self.doc_hash,
        }


# ---------------------------------------------------------------------------
# LibreOffice headless recompute
# ---------------------------------------------------------------------------


def recompute_xlsx(xlsx_path: str, timeout: int = 60) -> str:
    """Recompute all formulas in an .xlsx file using LibreOffice headless.

    Converts the .xlsx to .xlsx (round-trip) via LibreOffice, which forces
    a full recalculation of all formulas. The recomputed file replaces the
    original.

    Args:
        xlsx_path: Path to the .xlsx file to recompute.
        timeout: Maximum time in seconds for the LibreOffice subprocess.

    Returns:
        Path to the recomputed .xlsx file.

    Raises:
        LibreOfficeUnavailableError: If LibreOffice is not installed.
        RuntimeError: If the recompute subprocess fails.
    """
    soffice = _find_libreoffice()
    if not soffice:
        raise LibreOfficeUnavailableError(
            "recompute engine unavailable — install LibreOffice "
            "(soffice --headless) to enable xlsx_recompute oracle. "
            "Formula values CANNOT be verified without recompute."
        )

    xlsx_path = str(Path(xlsx_path).resolve())
    str(Path(xlsx_path).parent)
    input_name = Path(xlsx_path).name

    # LibreOffice cannot overwrite the input file — it needs a separate output dir.
    # We output to a temp dir, then move the recomputed file back to the original path.
    with tempfile.TemporaryDirectory() as tmp_out:
        with tempfile.TemporaryDirectory() as profile_dir:
            cmd = [
                soffice,
                "--headless",
                "--calc",
                f"-env:UserInstallation=file://{profile_dir}",
                "--convert-to",
                "xlsx",
                "--outdir",
                tmp_out,
                xlsx_path,
            ]

            try:
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=timeout,
                )
            except subprocess.TimeoutExpired:
                raise RuntimeError(
                    f"LibreOffice recompute timed out after {timeout}s for {xlsx_path}"
                )

            if result.returncode != 0:
                raise RuntimeError(
                    f"LibreOffice recompute failed (exit {result.returncode}): "
                    f"{result.stderr.strip()}"
                )

        # LibreOffice writes the recomputed file to tmp_out/input_name
        recomputed = os.path.join(tmp_out, input_name)
        if not os.path.exists(recomputed):
            raise RuntimeError(
                f"LibreOffice recompute did not produce output file: {recomputed}"
            )

        # Move the recomputed file back to the original location
        shutil.move(recomputed, xlsx_path)

    return xlsx_path


def read_recomputed_values(xlsx_path: str) -> dict[str, Any]:
    """Read recomputed cell values from an .xlsx file (data_only=True).

    Returns a dict mapping "Sheet!Cell" → computed value.
    Only cells with formulas will have computed values; non-formula cells
    return their literal value.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    values: dict[str, Any] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    key = f"{sheet_name}!{cell.coordinate}"
                    values[key] = cell.value
    return values


def read_formulas(xlsx_path: str) -> dict[str, str]:
    """Read formula text from an .xlsx file (data_only=False).

    Returns a dict mapping "Sheet!Cell" → formula string.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    formulas: dict[str, str] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if (
                    cell.value is not None
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")
                ):
                    key = f"{sheet_name}!{cell.coordinate}"
                    formulas[key] = cell.value
    return formulas


# ---------------------------------------------------------------------------
# xlsx_recompute oracle
# ---------------------------------------------------------------------------


def xlsx_recompute(
    xlsx_path: str,
    expected_values: dict[str, float],
    tolerance: float = 1e-6,
) -> bool:
    """Oracle: recompute xlsx → read back values → assert equal to expected.

    Args:
        xlsx_path: Path to the .xlsx file to recompute and verify.
        expected_values: Dict mapping "Sheet!Cell" → expected numeric value.
            These values should come from an INDEPENDENT calculation, not from
            the formula text.
        tolerance: Float comparison tolerance.

    Returns:
        True if all recomputed values match the expected values within tolerance.

    Raises:
        AssertionError if any value doesn't match.
        LibreOfficeUnavailableError if LibreOffice is not installed.
    """
    # 1. Recompute via LibreOffice
    recompute_xlsx(xlsx_path)

    # 2. Read back the recomputed values
    actual_values = read_recomputed_values(xlsx_path)

    # 3. Compare to expected (independent calc)
    mismatches: list[str] = []
    for cell_key, expected in expected_values.items():
        actual = actual_values.get(cell_key)
        if actual is None:
            mismatches.append(f"{cell_key}: MISSING (expected {expected})")
            continue
        try:
            actual_float = float(actual)
            if abs(actual_float - float(expected)) > tolerance:
                mismatches.append(
                    f"{cell_key}: expected {expected}, got {actual_float} "
                    f"(diff {abs(actual_float - float(expected))})"
                )
        except (TypeError, ValueError):
            mismatches.append(
                f"{cell_key}: expected {expected} (numeric), got {actual} (non-numeric)"
            )

    if mismatches:
        raise AssertionError(
            f"xlsx_recompute FAILED: {len(mismatches)} value(s) mismatch:\n  "
            + "\n  ".join(mismatches)
        )

    return True


# ---------------------------------------------------------------------------
# xlsx_structure_readback oracle
# ---------------------------------------------------------------------------


def xlsx_structure_readback(
    xlsx_path: str,
    expected_sheets: list[str] | None = None,
    expected_named_ranges: list[str] | None = None,
) -> bool:
    """Oracle: verify xlsx structure (sheets, named ranges, formats) survive round-trip.

    Args:
        xlsx_path: Path to the .xlsx file to verify.
        expected_sheets: List of sheet names that must be present.
        expected_named_ranges: List of named range names that must be present.

    Returns:
        True if all structural elements are present.

    Raises:
        AssertionError if any structural element is missing.
    """
    wb = openpyxl.load_workbook(xlsx_path)

    # Check sheets
    if expected_sheets:
        actual_sheets = wb.sheetnames
        missing_sheets = [s for s in expected_sheets if s not in actual_sheets]
        if missing_sheets:
            raise AssertionError(
                f"xlsx_structure_readback FAILED: missing sheets: {missing_sheets} "
                f"(actual: {actual_sheets})"
            )

    # Check named ranges
    if expected_named_ranges:
        actual_names = list(wb.defined_names)
        missing_names = [n for n in expected_named_ranges if n not in actual_names]
        if missing_names:
            raise AssertionError(
                f"xlsx_structure_readback FAILED: missing named ranges: {missing_names} "
                f"(actual: {actual_names})"
            )

    return True


# ---------------------------------------------------------------------------
# Independent calculation (numpy/pandas — never trust the writer alone)
# ---------------------------------------------------------------------------


def independent_calc_financial_model(
    revenue: float,
    costs: float,
) -> dict[str, float]:
    """Independent Python calculation of the financial model logic.

    This is the INDEPENDENT verification — it computes the same logic as the
    Excel formulas but using Python/numpy, so we can compare against LibreOffice's
    recompute. Never trust the writer alone.
    """
    gross_profit = revenue - costs
    margin = gross_profit / revenue if revenue != 0 else 0.0
    return {
        "Financial Model!B3": gross_profit,
        "Financial Model!B4": margin,
        "Summary!B1": gross_profit,
        "Summary!B2": margin,
    }


def independent_calc_pivot_aggregation(
    data: list[tuple[str, float]],
) -> dict[str, float]:
    """Independent Python calculation of pivot-style aggregation.

    Groups data by category and computes SUM and AVERAGE per group.
    """
    from collections import defaultdict

    sums: dict[str, float] = defaultdict(float)
    counts: dict[str, int] = defaultdict(int)

    for category, value in data:
        sums[category] += value
        counts[category] += 1

    result: dict[str, float] = {}
    for category in sorted(sums.keys()):
        result[f"Pivot!{category}_sum"] = sums[category]
        result[f"Pivot!{category}_avg"] = sums[category] / counts[category]

    return result


# ---------------------------------------------------------------------------
# Excel pipeline with trust stack integration
# ---------------------------------------------------------------------------


def excel_pipeline(
    input_path: str,
    output_path: str,
    edits: list[dict[str, Any]],
    expected_values: dict[str, float],
    expected_sheets: list[str] | None = None,
    expected_named_ranges: list[str] | None = None,
    private_key: Any = None,
    author: str = "Kairo Excel",
) -> ExcelResult:
    """Run the Excel pipeline with recompute verification + trust stack.

    1. Load the input workbook.
    2. Apply the specified edits (formulas + values).
    3. Save to output_path.
    4. Recompute via LibreOffice headless.
    5. Read back recomputed values and verify against independent calc.
    6. Verify structure (sheets, named ranges).
    7. Emit Ed25519 audit log + zero-egress report (if private_key provided).

    Args:
        input_path: Path to the input .xlsx file.
        output_path: Path where the recomputed .xlsx will be saved.
        edits: List of edit dicts with 'sheet', 'cell', 'value', 'rationale'.
        expected_values: Dict of "Sheet!Cell" → expected numeric value
            from independent calculation.
        expected_sheets: List of sheet names that must survive round-trip.
        expected_named_ranges: List of named ranges that must survive.
        private_key: Optional Ed25519 private key for audit + egress report.
        author: Author name for audit log.

    Returns:
        ExcelResult with recompute verification status.
    """
    input_path = str(Path(input_path).resolve())
    output_path = str(Path(output_path).resolve())

    if not os.path.exists(input_path):
        return ExcelResult(ok=False, error=f"Input file not found: {input_path}")

    # 1. Load workbook
    wb = openpyxl.load_workbook(input_path)

    # 2. Apply edits
    applied: list[ExcelEdit] = []
    for edit in edits:
        sheet_name = edit.get("sheet", "")
        cell_ref = edit.get("cell", "")
        new_value = edit.get("value", "")
        rationale = edit.get("rationale", "")

        if sheet_name not in wb.sheetnames:
            return ExcelResult(
                ok=False,
                error=f"Sheet '{sheet_name}' not found in workbook. "
                f"Available: {wb.sheetnames}",
            )

        ws = wb[sheet_name]
        old_value = ws[cell_ref].value
        ws[cell_ref] = new_value

        applied.append(
            ExcelEdit(
                sheet=sheet_name,
                cell=cell_ref,
                old_value=old_value,
                new_value=new_value,
                formula=str(new_value)
                if isinstance(new_value, str) and new_value.startswith("=")
                else "",
                rationale=rationale,
            )
        )

    # 3. Save
    wb.save(output_path)

    # 4. Compute doc hash
    with open(output_path, "rb") as f:
        doc_hash = hashlib.sha256(f.read()).hexdigest()

    # 5. Recompute via LibreOffice
    try:
        recompute_xlsx(output_path)
    except LibreOfficeUnavailableError as e:
        return ExcelResult(
            ok=False,
            error=str(e),
            applied_edits=applied,
            doc_hash=doc_hash,
        )
    except RuntimeError as e:
        return ExcelResult(
            ok=False,
            error=f"Recompute failed: {e}",
            applied_edits=applied,
            doc_hash=doc_hash,
        )

    # 6. Read back recomputed values
    actual_values = read_recomputed_values(output_path)

    # 7. Verify against independent calc
    recompute_verified = True
    mismatches: list[str] = []
    for cell_key, expected in expected_values.items():
        actual = actual_values.get(cell_key)
        if actual is None:
            mismatches.append(f"{cell_key}: MISSING")
            recompute_verified = False
            continue
        try:
            actual_float = float(actual)
            if abs(actual_float - float(expected)) > 1e-6:
                mismatches.append(
                    f"{cell_key}: expected {expected}, got {actual_float}"
                )
                recompute_verified = False
        except (TypeError, ValueError):
            mismatches.append(f"{cell_key}: expected {expected}, got {actual}")
            recompute_verified = False

    # 8. Verify structure
    if expected_sheets or expected_named_ranges:
        try:
            xlsx_structure_readback(output_path, expected_sheets, expected_named_ranges)
        except AssertionError as e:
            return ExcelResult(
                ok=False,
                error=str(e),
                applied_edits=applied,
                recomputed_values=actual_values,
                independent_calc=expected_values,
                recompute_verified=False,
                doc_hash=doc_hash,
            )

    # 9. Emit audit log + egress report (if key provided)
    audit_log_json = ""
    egress_report_json = ""
    if private_key is not None:
        from kairo.oracles.ed25519_audit_log import Ed25519AuditLog
        from kairo.oracles.zero_egress_report import generate_zero_egress_report

        audit = Ed25519AuditLog(private_key)
        audit.log_run_started(doc_hash=doc_hash, playbook_id="excel_pipeline")

        for edit in applied:
            audit.log_edit(
                doc_hash=doc_hash,
                clause_id=f"{edit.sheet}!{edit.cell}",
                clause_label=f"Excel edit: {edit.sheet}!{edit.cell}",
                old_text=str(edit.old_value),
                new_text=str(edit.new_value),
                citation=edit.formula or "direct value",
                rationale=edit.rationale,
            )

        audit.log_run_completed(
            doc_hash=doc_hash,
            total_edits=len(applied),
            total_flagged=0,
            injection_detected=False,
        )

        audit_log_json = audit.to_json()

        egress_report = generate_zero_egress_report(
            doc_hash=doc_hash,
            playbook_id="excel_pipeline",
            total_edits=len(applied),
            total_flagged=0,
            injection_detected=False,
            audit_log_json=audit_log_json,
            private_key=private_key,
        )
        egress_report_json = egress_report.to_json()

    return ExcelResult(
        ok=recompute_verified,
        output_path=output_path,
        applied_edits=applied,
        recomputed_values=actual_values,
        independent_calc=expected_values,
        recompute_verified=recompute_verified,
        error="; ".join(mismatches) if mismatches else "",
        audit_log_json=audit_log_json,
        egress_report_json=egress_report_json,
        doc_hash=doc_hash,
    )
