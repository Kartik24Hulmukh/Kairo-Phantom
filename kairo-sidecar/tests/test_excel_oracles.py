# PROVENANCE: original | clean-room Excel oracle tests per specs/VERIFICATION_ORACLES.md
"""Excel domain oracle tests — xlsx_recompute + xlsx_structure_readback + kill-proofs.

Tests verify:
  1. xlsx_recompute: LibreOffice recalc → read back values → assert equal to
     independent Python calc.
  2. xlsx_structure_readback: sheets / named ranges survive round-trip.
  3. Kill-proofs: break recompute → FAILS; drop a sheet → FAILS.
  4. Honest degradation: LibreOffice missing → FAIL LOUD.
  5. >=3 gauntlet scenarios: financial model, pivot aggregation, wrong formula.
  6. Trust stack integration: audit log + egress report + sealed mode.

All tests run fully offline. No mocks on production paths.
"""

from __future__ import annotations

import json
import os
import sys
import tempfile

import openpyxl
import pytest
from cryptography.hazmat.primitives.asymmetric import ed25519

# Ensure repo root is on sys.path
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from kairo.excel.engine import (
    recompute_xlsx,
    xlsx_recompute,
    xlsx_structure_readback,
    excel_pipeline,
    LibreOfficeUnavailableError,
    independent_calc_financial_model,
    independent_calc_pivot_aggregation,
)

# Fixture paths
_EXCEL_DIR = os.path.join(_REPO_ROOT, "fixtures", "excel")


# ---------------------------------------------------------------------------
# Helper: check if LibreOffice is available
# ---------------------------------------------------------------------------


def _libreoffice_available() -> bool:
    """Check if LibreOffice is available on this system."""
    import shutil

    for name in ("soffice", "libreoffice"):
        if shutil.which(name):
            return True
    common = ["/usr/bin/soffice", "/usr/bin/libreoffice", "/usr/local/bin/soffice"]
    return any(os.path.isfile(p) for p in common)


_HAS_LIBREOFFICE = _libreoffice_available()


# ---------------------------------------------------------------------------
# xlsx_recompute oracle tests
# ---------------------------------------------------------------------------


class TestXlsxRecompute:
    """xlsx_recompute oracle: recompute → read back → verify vs independent calc."""

    def test_financial_model_recompute(self):
        """Financial model: LibreOffice recomputes formulas, values match independent calc."""
        if not _HAS_LIBREOFFICE:
            pytest.fail("LibreOffice not available — cannot test xlsx_recompute")

        with tempfile.TemporaryDirectory() as tmp:
            # Copy fixture
            import shutil

            src = os.path.join(_EXCEL_DIR, "s01_financial_model.xlsx")
            dst = os.path.join(tmp, "financial_model.xlsx")
            shutil.copy(src, dst)

            # Apply edits (update revenue and costs)
            wb = openpyxl.load_workbook(dst)
            wb["Financial Model"]["B1"] = 120000
            wb["Financial Model"]["B2"] = 70000
            wb.save(dst)

            # Independent calc
            expected = independent_calc_financial_model(120000, 70000)

            # Recompute and verify
            result = xlsx_recompute(dst, expected)
            assert result is True, "xlsx_recompute should pass for correct formulas"

    def test_pivot_aggregation_recompute(self):
        """Pivot aggregation: SUMIF/AVERAGEIF recomputed values match independent calc."""
        if not _HAS_LIBREOFFICE:
            pytest.fail("LibreOffice not available — cannot test xlsx_recompute")

        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s02_pivot_aggregation.xlsx")
            dst = os.path.join(tmp, "pivot.xlsx")
            shutil.copy(src, dst)

            # Read ground truth
            with open(os.path.join(_EXCEL_DIR, "s02_ground_truth.json")) as f:
                expected = json.load(f)

            # Convert string values to float
            expected = {k: float(v) for k, v in expected.items()}

            # Recompute and verify
            result = xlsx_recompute(dst, expected)
            assert result is True, "xlsx_recompute should pass for SUMIF/AVERAGEIF"

    def test_kill_proof_wrong_formula(self):
        """Kill-proof: a deliberately wrong formula MUST be caught by the oracle."""
        if not _HAS_LIBREOFFICE:
            pytest.fail("LibreOffice not available — cannot test xlsx_recompute")

        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s03_wrong_formula.xlsx")
            dst = os.path.join(tmp, "wrong.xlsx")
            shutil.copy(src, dst)

            # Ground truth says B4 should be 200 (=B1*B2) but formula is =B1+B2=30
            with open(os.path.join(_EXCEL_DIR, "s03_ground_truth.json")) as f:
                expected = json.load(f)
            expected = {k: float(v) for k, v in expected.items()}

            # The oracle MUST fail — the formula gives 30, not 200
            with pytest.raises(AssertionError, match="B4"):
                xlsx_recompute(dst, expected)

    def test_kill_proof_formula_string_not_value(self):
        """Kill-proof: if recompute is skipped, formula cells have no cached value (None).

        The oracle MUST fail when given expected values that don't match the
        uncomputed (None) cell values. This proves the oracle catches the case
        where someone presents formula strings without recompute.
        """
        if not _HAS_LIBREOFFICE:
            pytest.fail("LibreOffice not available — cannot test xlsx_recompute")

        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s01_financial_model.xlsx")
            dst = os.path.join(tmp, "test.xlsx")
            shutil.copy(src, dst)

            # Read WITHOUT recompute — formula cells will have None as their
            # cached value (data_only=True returns None for uncomputed formulas)
            wb = openpyxl.load_workbook(dst, data_only=True)
            b3_val = wb["Financial Model"]["B3"].value
            # B3 has formula =B1-B2, but without recompute, data_only gives None
            assert b3_val is None, (
                f"Expected None for uncomputed formula cell B3, got {b3_val} — "
                f"this means the oracle would NOT catch missing recompute"
            )

            # Now verify the oracle WOULD fail: call xlsx_recompute with expected
            # values that don't match the uncomputed (None) cells
            # We monkeypatch recompute_xlsx to be a no-op (simulating broken recompute)
            import kairo.excel.engine as engine_mod

            original_recompute = engine_mod.recompute_xlsx

            def fake_recompute(path, timeout=60):
                # Don't actually recompute — just return the path
                return path

            engine_mod.recompute_xlsx = fake_recompute
            try:
                with pytest.raises(AssertionError, match="B3"):
                    xlsx_recompute(dst, {"Financial Model!B3": 50000.0})
            finally:
                engine_mod.recompute_xlsx = original_recompute


# ---------------------------------------------------------------------------
# xlsx_structure_readback oracle tests
# ---------------------------------------------------------------------------


class TestXlsxStructureReadback:
    """xlsx_structure_readback oracle: sheets / named ranges survive round-trip."""

    def test_structure_survives_roundtrip(self):
        """Sheets and named ranges survive a save/load round-trip."""
        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s01_financial_model.xlsx")
            dst = os.path.join(tmp, "test.xlsx")
            shutil.copy(src, dst)

            result = xlsx_structure_readback(
                dst,
                expected_sheets=["Financial Model", "Summary"],
                expected_named_ranges=["TotalRevenue"],
            )
            assert result is True

    def test_kill_proof_missing_sheet(self):
        """Kill-proof: dropping a sheet MUST fail the structure oracle."""
        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s01_financial_model.xlsx")
            dst = os.path.join(tmp, "test.xlsx")
            shutil.copy(src, dst)

            # Remove a sheet
            wb = openpyxl.load_workbook(dst)
            del wb["Summary"]
            wb.save(dst)

            with pytest.raises(AssertionError, match="missing sheets"):
                xlsx_structure_readback(dst, expected_sheets=["Financial Model", "Summary"])

    def test_kill_proof_missing_named_range(self):
        """Kill-proof: dropping a named range MUST fail the structure oracle."""
        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s01_financial_model.xlsx")
            dst = os.path.join(tmp, "test.xlsx")
            shutil.copy(src, dst)

            # Remove a named range
            wb = openpyxl.load_workbook(dst)
            del wb.defined_names["TotalRevenue"]
            wb.save(dst)

            with pytest.raises(AssertionError, match="missing named ranges"):
                xlsx_structure_readback(dst, expected_named_ranges=["TotalRevenue"])


# ---------------------------------------------------------------------------
# Honest degradation tests
# ---------------------------------------------------------------------------


class TestHonestDegradation:
    """LibreOffice missing → FAIL LOUD. Never present unverified formula strings."""

    def test_libreoffice_unavailable_raises(self, monkeypatch):
        """When LibreOffice is not installed, recompute raises LibreOfficeUnavailableError."""
        # Monkeypatch _find_libreoffice to return None
        import kairo.excel.engine as engine

        monkeypatch.setattr(engine, "_find_libreoffice", lambda: None)

        with pytest.raises(LibreOfficeUnavailableError, match="recompute engine unavailable"):
            recompute_xlsx("/tmp/nonexistent.xlsx")

    def test_pipeline_fails_loud_without_libreoffice(self, monkeypatch):
        """The pipeline returns ok=False with a clear error when LibreOffice is missing."""
        import kairo.excel.engine as engine

        monkeypatch.setattr(engine, "_find_libreoffice", lambda: None)

        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s01_financial_model.xlsx")
            dst = os.path.join(tmp, "output.xlsx")
            shutil.copy(src, dst)

            result = excel_pipeline(
                input_path=src,
                output_path=dst,
                edits=[
                    {"sheet": "Financial Model", "cell": "B1", "value": 120000, "rationale": "test"}
                ],
                expected_values={"Financial Model!B3": 50000.0},
            )
            assert not result.ok, "Pipeline should fail when LibreOffice is missing"
            assert "unavailable" in result.error.lower() or "recompute" in result.error.lower()


# ---------------------------------------------------------------------------
# Excel pipeline with trust stack integration
# ---------------------------------------------------------------------------


class TestExcelPipelineTrustStack:
    """The Excel pipeline integrates with the existing trust stack."""

    def test_pipeline_emits_audit_log(self):
        """The pipeline emits an Ed25519-signed audit log."""
        if not _HAS_LIBREOFFICE:
            pytest.fail("LibreOffice not available")

        private_key = ed25519.Ed25519PrivateKey.generate()

        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s01_financial_model.xlsx")
            dst = os.path.join(tmp, "output.xlsx")
            shutil.copy(src, dst)

            expected = independent_calc_financial_model(120000, 70000)

            result = excel_pipeline(
                input_path=src,
                output_path=dst,
                edits=[
                    {
                        "sheet": "Financial Model",
                        "cell": "B1",
                        "value": 120000,
                        "rationale": "Update revenue",
                    },
                    {
                        "sheet": "Financial Model",
                        "cell": "B2",
                        "value": 70000,
                        "rationale": "Update costs",
                    },
                ],
                expected_values=expected,
                expected_sheets=["Financial Model", "Summary"],
                expected_named_ranges=["TotalRevenue"],
                private_key=private_key,
            )

            assert result.ok, f"Pipeline failed: {result.error}"
            assert result.audit_log_json, "No audit log emitted"
            assert result.egress_report_json, "No egress report emitted"

            # Verify audit log
            from kairo.oracles.ed25519_audit_log import Ed25519AuditLog

            entries = Ed25519AuditLog.entries_from_json(result.audit_log_json)
            public_key = private_key.public_key()
            assert Ed25519AuditLog.verify_chain(entries, public_key), "Audit log chain invalid"

    def test_pipeline_egress_report_verified(self):
        """The egress report signature is valid."""
        if not _HAS_LIBREOFFICE:
            pytest.fail("LibreOffice not available")

        private_key = ed25519.Ed25519PrivateKey.generate()

        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s01_financial_model.xlsx")
            dst = os.path.join(tmp, "output.xlsx")
            shutil.copy(src, dst)

            expected = independent_calc_financial_model(120000, 70000)

            result = excel_pipeline(
                input_path=src,
                output_path=dst,
                edits=[
                    {"sheet": "Financial Model", "cell": "B1", "value": 120000, "rationale": "test"}
                ],
                expected_values=expected,
                private_key=private_key,
            )

            from kairo.oracles.zero_egress_report import report_from_json, verify_zero_egress_report

            report = report_from_json(result.egress_report_json)
            public_key = private_key.public_key()
            assert verify_zero_egress_report(report, public_key), "Egress report signature invalid"


# ---------------------------------------------------------------------------
# Gauntlet scenarios (>=3)
# ---------------------------------------------------------------------------


class TestExcelGauntletScenarios:
    """>=3 gauntlet scenarios: financial model, pivot aggregation, wrong formula."""

    def test_scenario_1_financial_model(self):
        """Scenario 1: financial model with cross-sheet deps + named ranges."""
        if not _HAS_LIBREOFFICE:
            pytest.fail("LibreOffice not available")

        private_key = ed25519.Ed25519PrivateKey.generate()

        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s01_financial_model.xlsx")
            dst = os.path.join(tmp, "output.xlsx")
            shutil.copy(src, dst)

            expected = independent_calc_financial_model(120000, 70000)

            result = excel_pipeline(
                input_path=src,
                output_path=dst,
                edits=[
                    {
                        "sheet": "Financial Model",
                        "cell": "B1",
                        "value": 120000,
                        "rationale": "Update revenue",
                    },
                    {
                        "sheet": "Financial Model",
                        "cell": "B2",
                        "value": 70000,
                        "rationale": "Update costs",
                    },
                ],
                expected_values=expected,
                expected_sheets=["Financial Model", "Summary"],
                expected_named_ranges=["TotalRevenue"],
                private_key=private_key,
            )

            assert result.ok, f"Financial model scenario failed: {result.error}"
            assert result.recompute_verified, "Recompute not verified"
            assert len(result.applied_edits) == 2

    def test_scenario_2_pivot_aggregation(self):
        """Scenario 2: pivot-style aggregation with SUMIF/AVERAGEIF."""
        if not _HAS_LIBREOFFICE:
            pytest.fail("LibreOffice not available")

        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s02_pivot_aggregation.xlsx")
            dst = os.path.join(tmp, "output.xlsx")
            shutil.copy(src, dst)

            with open(os.path.join(_EXCEL_DIR, "s02_ground_truth.json")) as f:
                expected = {k: float(v) for k, v in json.load(f).items()}

            result = excel_pipeline(
                input_path=src,
                output_path=dst,
                edits=[],
                expected_values=expected,
                expected_sheets=["Data", "Pivot"],
            )

            assert result.ok, f"Pivot aggregation scenario failed: {result.error}"
            assert result.recompute_verified, "Recompute not verified"

    def test_scenario_3_wrong_formula_caught(self):
        """Scenario 3: deliberately wrong formula — oracle MUST catch it."""
        if not _HAS_LIBREOFFICE:
            pytest.fail("LibreOffice not available")

        with tempfile.TemporaryDirectory() as tmp:
            import shutil

            src = os.path.join(_EXCEL_DIR, "s03_wrong_formula.xlsx")
            dst = os.path.join(tmp, "output.xlsx")
            shutil.copy(src, dst)

            with open(os.path.join(_EXCEL_DIR, "s03_ground_truth.json")) as f:
                expected = {k: float(v) for k, v in json.load(f).items()}

            result = excel_pipeline(
                input_path=src,
                output_path=dst,
                edits=[],
                expected_values=expected,
                expected_sheets=["Calculation"],
            )

            # The pipeline should FAIL — B4 gives 30 but expected 200
            assert not result.ok, "Wrong formula was NOT caught by the oracle"
            assert "B4" in result.error, f"Error should mention B4: {result.error}"


# ---------------------------------------------------------------------------
# Independent calculation tests
# ---------------------------------------------------------------------------


class TestIndependentCalc:
    """Independent Python calculations used for oracle verification."""

    def test_financial_model_calc(self):
        """Independent calc of financial model produces correct values."""
        result = independent_calc_financial_model(120000, 70000)
        assert result["Financial Model!B3"] == 50000.0
        assert abs(result["Financial Model!B4"] - 50000.0 / 120000.0) < 1e-10

    def test_pivot_aggregation_calc(self):
        """Independent calc of pivot aggregation produces correct values."""
        data = [
            ("Electronics", 1500),
            ("Electronics", 2300),
            ("Electronics", 1800),
            ("Books", 450),
            ("Books", 620),
            ("Clothing", 890),
            ("Clothing", 1200),
            ("Clothing", 750),
        ]
        result = independent_calc_pivot_aggregation(data)
        assert result["Pivot!Electronics_sum"] == 5600.0
        assert abs(result["Pivot!Electronics_avg"] - 5600.0 / 3) < 1e-10
        assert result["Pivot!Books_sum"] == 1070.0
        assert result["Pivot!Clothing_sum"] == 2840.0
